"""Tests for smart Excel parser (header detection, fuzzy matching, heuristics, fallback)."""

import io

import pandas as pd
import pytest
from openpyxl import Workbook

from app.parser_excel import ParseError, build_dataframe_from_columns, load_excel, parse_excel


def _save_wb(wb, tmp_path, name="test.xlsx"):
    path = tmp_path / name
    wb.save(str(path))
    return path


# ── Header not in first row ──────────────────────────────


def test_parse_xlsx_header_not_first_row(tmp_path):
    """Parser finds header row even when it's not the first row (e.g. row 5)."""
    wb = Workbook()
    ws = wb.active

    # Rows 1-4: junk / decorative content
    ws.cell(row=1, column=1, value="Компания ООО «МетизТорг»")
    ws.cell(row=2, column=1, value="Заявка на закупку №123 от 01.01.2025")
    ws.cell(row=3, column=1, value="")
    ws.cell(row=4, column=1, value="Примечание: доставка до склада")

    # Row 5: actual header
    ws.cell(row=5, column=1, value="Код")
    ws.cell(row=5, column=2, value="Номенклатура")
    ws.cell(row=5, column=3, value="Кол-во")

    # Rows 6-8: data
    ws.cell(row=6, column=1, value="001")
    ws.cell(row=6, column=2, value="Болт М12x80")
    ws.cell(row=6, column=3, value=100)

    ws.cell(row=7, column=1, value="002")
    ws.cell(row=7, column=2, value="Гайка М16")
    ws.cell(row=7, column=3, value=200)

    ws.cell(row=8, column=1, value="003")
    ws.cell(row=8, column=2, value="Шайба 10")
    ws.cell(row=8, column=3, value=50)

    path = _save_wb(wb, tmp_path)
    df = load_excel(path)

    assert len(df) == 3
    assert list(df.columns) == ["code", "name", "qty", "uom"]
    assert df.iloc[0]["code"] == "001"
    assert df.iloc[0]["name"] == "Болт М12x80"
    assert df.iloc[0]["qty"] == 100
    assert df.iloc[1]["name"] == "Гайка М16"
    assert df.iloc[2]["qty"] == 50


# ── Ignores note block after table ──────────────────────


def test_parse_xlsx_ignores_note_block(tmp_path):
    """Parser skips rows without a name value (e.g. notes/totals after table)."""
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Артикул")
    ws.cell(row=1, column=2, value="Наименование")
    ws.cell(row=1, column=3, value="Заказ")

    ws.cell(row=2, column=1, value="A01")
    ws.cell(row=2, column=2, value="Болт М8x40")
    ws.cell(row=2, column=3, value=50)

    ws.cell(row=3, column=1, value="A02")
    ws.cell(row=3, column=2, value="Гайка М8")
    ws.cell(row=3, column=3, value=100)

    ws.cell(row=4, column=1, value="")
    ws.cell(row=4, column=2, value="")
    ws.cell(row=4, column=3, value=150)

    ws.cell(row=5, column=1, value="Примечание:")
    ws.cell(row=5, column=2, value="")
    ws.cell(row=5, column=3, value="")

    path = _save_wb(wb, tmp_path)
    df = load_excel(path)

    assert len(df) == 2
    assert df.iloc[0]["name"] == "Болт М8x40"
    assert df.iloc[1]["name"] == "Гайка М8"


# ── .xls rejection via web upload ────────────────────────


def test_upload_xls_rejected(tmp_path):
    """Web upload endpoint rejects .xls files with a clear error message."""
    import app.cache as cache_mod

    upload_dir = tmp_path / "uploads"
    cache_dir = tmp_path / "cache"
    cache_mod.UPLOAD_DIR = upload_dir
    cache_mod.CACHE_DIR = cache_dir

    from fastapi.testclient import TestClient

    from app.main import app

    client = TestClient(app)

    buf = io.BytesIO(b"fake xls content")
    resp = client.post(
        "/upload",
        files={"file": ("data.xls", buf, "application/vnd.ms-excel")},
    )
    assert resp.status_code == 400
    assert ".xls" in resp.text
    assert "xlsx" in resp.text.lower()


# ── Synonym variations ───────────────────────────────────


def test_parse_xlsx_synonym_variations(tmp_path):
    """Parser recognizes synonym variations for column headers."""
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Артикул")
    ws.cell(row=1, column=2, value="Товар")
    ws.cell(row=1, column=3, value="Количество")

    ws.cell(row=2, column=1, value="X100")
    ws.cell(row=2, column=2, value="Шуруп 5x40")
    ws.cell(row=2, column=3, value=300)

    path = _save_wb(wb, tmp_path)
    df = load_excel(path)

    assert len(df) == 1
    assert df.iloc[0]["code"] == "X100"
    assert df.iloc[0]["name"] == "Шуруп 5x40"
    assert df.iloc[0]["qty"] == 300


# ── Auto-detect with standard headers (parse_excel) ─────


def test_auto_detect_standard_headers(tmp_path):
    """parse_excel() finds NAME and QTY with standard headers, needs_manual_selection=False."""
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Код")
    ws.cell(row=1, column=2, value="Номенклатура")
    ws.cell(row=1, column=3, value="Заказ")

    ws.cell(row=2, column=1, value="001")
    ws.cell(row=2, column=2, value="Болт М12x80")
    ws.cell(row=2, column=3, value=100)

    ws.cell(row=3, column=1, value="002")
    ws.cell(row=3, column=2, value="Гайка М16")
    ws.cell(row=3, column=3, value=50)

    path = _save_wb(wb, tmp_path)
    result = parse_excel(path)

    assert not result.needs_manual_selection
    assert result.df is not None
    assert len(result.df) == 2
    assert result.detected.name_idx is not None
    assert result.detected.qty_idx is not None
    assert result.detected.score >= 2


# ── Auto-detect by content (no explicit qty header) ──────


def test_auto_detect_by_content(tmp_path):
    """File with NAME header but no explicit QTY header — numeric column detected via heuristic."""
    wb = Workbook()
    ws = wb.active

    # Header with recognizable name but no qty synonym
    ws.cell(row=1, column=1, value="Артикул")
    ws.cell(row=1, column=2, value="Наименование")
    ws.cell(row=1, column=3, value="Примечание")  # Not a qty synonym

    # Data: column 3 is actually numeric (quantities)
    for i in range(2, 12):
        ws.cell(row=i, column=1, value=f"A{i:02d}")
        ws.cell(row=i, column=2, value=f"Товар {i}")
        ws.cell(row=i, column=3, value=i * 10)

    path = _save_wb(wb, tmp_path)
    result = parse_excel(path)

    assert not result.needs_manual_selection
    assert result.df is not None
    assert len(result.df) == 10
    # QTY should have been detected via content heuristic
    assert result.detected.qty_idx is not None
    assert result.df.iloc[0]["qty"] == 20


# ── Fallback: unrecognizable headers ─────────────────────


def test_fallback_returns_needs_manual_selection(tmp_path):
    """File with completely unrecognizable headers triggers fallback."""
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Alpha")
    ws.cell(row=1, column=2, value="Beta")
    ws.cell(row=1, column=3, value="Gamma")

    ws.cell(row=2, column=1, value="aaa")
    ws.cell(row=2, column=2, value="bbb")
    ws.cell(row=2, column=3, value="ccc")

    ws.cell(row=3, column=1, value="ddd")
    ws.cell(row=3, column=2, value="eee")
    ws.cell(row=3, column=3, value="fff")

    path = _save_wb(wb, tmp_path)
    result = parse_excel(path)

    assert result.needs_manual_selection
    assert result.df is None
    assert result.raw_values is not None
    assert len(result.raw_values) >= 3


# ── build_dataframe_from_columns ─────────────────────────


def test_build_dataframe_from_columns():
    """Directly test the helper that builds DataFrame from raw values + column indices."""
    values_2d = [
        ["Код", "Наименование", "Кол-во"],  # row 0 = header
        ["001", "Болт M10", 50],
        ["002", "Гайка M12", 100],
        ["", "", ""],  # empty row — should be skipped
    ]
    df = build_dataframe_from_columns(values_2d, header_idx=0, name_idx=1, qty_idx=2, code_idx=0)

    assert len(df) == 2
    assert list(df.columns) == ["code", "name", "qty", "uom"]
    assert df.iloc[0]["code"] == "001"
    assert df.iloc[0]["name"] == "Болт M10"
    assert df.iloc[0]["qty"] == 50
    assert df.iloc[1]["qty"] == 100


# ── ParseError on missing columns via load_excel ─────────


def test_parse_error_on_missing_columns(tmp_path):
    """load_excel() raises ParseError when columns cannot be determined."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Alpha")
    ws.cell(row=1, column=2, value="Beta")
    ws.cell(row=2, column=1, value="aaa")
    ws.cell(row=2, column=2, value="bbb")

    path = _save_wb(wb, tmp_path)
    with pytest.raises(ParseError):
        load_excel(path)
