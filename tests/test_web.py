import io
import re

import pandas as pd
import pytest
from fastapi.testclient import TestClient


@pytest.fixture(autouse=True)
def _set_dirs(tmp_path, monkeypatch):
    upload_dir = tmp_path / "uploads"
    cache_dir = tmp_path / "cache"
    monkeypatch.setenv("OTDELZAKUP_UPLOAD_DIR", str(upload_dir))
    monkeypatch.setenv("OTDELZAKUP_CACHE_DIR", str(cache_dir))
    import app.cache as cache_mod

    cache_mod.UPLOAD_DIR = upload_dir
    cache_mod.CACHE_DIR = cache_dir

    # Isolate DB per test
    db_path = tmp_path / "test.db"
    monkeypatch.setenv("OTDELZAKUP_DB_PATH", str(db_path))
    import app.database as db_mod
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_mod.DB_PATH = db_path
    db_mod.engine = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
    db_mod.SessionLocal = sessionmaker(bind=db_mod.engine, autoflush=False, expire_on_commit=False)
    db_mod.init_db()


@pytest.fixture()
def client():
    from app.main import app

    return TestClient(app)


def _make_xlsx(rows: list[dict]) -> io.BytesIO:
    """Create an in-memory .xlsx file from a list of dicts."""
    buf = io.BytesIO()
    pd.DataFrame(rows).to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf


def _upload_file(client, rows, filename="test.xlsx"):
    """Helper: upload an xlsx and return the response."""
    xlsx = _make_xlsx(rows)
    return client.post(
        "/upload",
        files={"file": (filename, xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def _extract_file_id(html: str) -> str:
    m = re.search(r'name="file_id"\s+value="([^"]+)"', html)
    assert m, "file_id hidden input not found"
    return m.group(1)


# ── 1. GET / ──────────────────────────────────────────────


def test_get_root_ok(client):
    resp = client.get("/")
    assert resp.status_code == 200
    html = resp.text
    assert "<form" in html
    assert 'name="file"' in html


# ── 2. POST /upload — happy path ─────────────────────────


def test_upload_xlsx_ok(client):
    data = [
        {"Код": "001", "Номенклатура": "Болт M12", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Гайка М16", "Заказ": 20},
    ]
    xlsx = _make_xlsx(data)

    resp = client.post(
        "/upload",
        files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    html = resp.text
    assert "test.xlsx" in html
    assert "2" in html  # total rows
    assert "Болт M12" in html
    assert "Гайка М16" in html


# ── 3. Non-.xlsx rejected ────────────────────────────────


def test_upload_not_xlsx_rejected(client):
    buf = io.BytesIO(b"just some text")
    resp = client.post(
        "/upload",
        files={"file": ("notes.txt", buf, "text/plain")},
    )
    assert resp.status_code == 400
    assert ".xlsx" in resp.text


# ── 4. Preview limited to 200 rows ──────────────────────


def test_upload_limits_preview_to_200_rows(client):
    rows = [{"Код": str(i), "Номенклатура": f"Товар_{i}", "Заказ": i} for i in range(250)]
    xlsx = _make_xlsx(rows)

    resp = client.post(
        "/upload",
        files={"file": ("big.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    html = resp.text
    assert "250" in html
    assert "Товар_199" in html
    assert "Товар_240" not in html


# ── 5. Upload page shows checkboxes ─────────────────────


def test_upload_shows_checkboxes(client):
    rows = [{"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 5}]
    resp = _upload_file(client, rows)
    assert resp.status_code == 200
    html = resp.text
    assert 'name="fields"' in html
    assert 'value="diameter"' in html
    assert 'value="strength"' in html
    assert "file_id" in html


# ── 6. Transform with selected fields ───────────────────


def test_transform_with_selected_fields(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Гайка М16 10.9 DIN 934", "Заказ": 20},
    ]
    resp = _upload_file(client, rows)
    assert resp.status_code == 200
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["diameter", "strength"]},
    )
    assert resp2.status_code == 200
    html = resp2.text
    assert "Исходная таблица" in html
    assert "Преобразованная таблица" in html
    assert "Диаметр" in html
    assert "Класс прочности" in html
    assert "M12" in html
    assert "8.8" in html
    assert "M16" in html
    assert "10.9" in html


# ── 7. Transform without fields ─────────────────────────


def test_transform_without_fields(client):
    rows = [{"Код": "001", "Номенклатура": "Болт M12x80 8.8", "Заказ": 5}]
    resp = _upload_file(client, rows)
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id},
    )
    assert resp2.status_code == 200
    html = resp2.text
    assert "Диаметр" not in html
    assert "Класс прочности" not in html
    assert "M12x80" in html


# ── 8. Extractors: basic metiz parsing ──────────────────


def test_extractors_basic_metiz():
    from app.extractors import (
        extract_coating,
        extract_diameter,
        extract_gost,
        extract_length,
        extract_size,
        extract_strength,
        extract_tail_code,
    )

    text = "Болт М12-6gx150.88.016 ГОСТ 7798-70"

    assert extract_diameter(text) == "M12"
    assert extract_length(text) == "150"
    assert extract_size(text) == "M12x150"
    assert extract_strength(text) == "8.8"
    assert extract_coating(text) == "цинк"
    assert "7798-70" in extract_gost(text)
    assert extract_tail_code(text) == ".88.016"


# ── 8a. Improved extractors ──────────────────────────


def test_normalize_m_and_x():
    from app.extractors import preprocess

    s = preprocess("Болт М12-6gx150")
    assert "m12" in s
    assert "x150" in s


def test_extract_metric_size():
    from app.extractors import (
        extract_coating,
        extract_diameter,
        extract_gost,
        extract_length,
        extract_size,
        extract_strength,
    )

    text = "Болт М12-6gx150.88.016 ГОСТ7798-70"
    assert extract_size(text) == "M12x150"
    assert extract_diameter(text) == "M12"
    assert extract_length(text) == "150"
    assert extract_strength(text) == "8.8"
    assert extract_coating(text) == "цинк"
    assert extract_gost(text) == "ГОСТ 7798-70"


def test_extract_screw_size():
    from app.extractors import (
        extract_item_type,
        extract_length,
        extract_screw_diameter,
        extract_size,
    )

    text = "Саморез 4,2х16 по металлу"
    assert extract_size(text) == "4.2x16"
    assert extract_screw_diameter(text) == "4.2"
    assert extract_length(text) == "16"
    assert extract_item_type(text) == "саморез"


def test_extract_coating_stainless():
    from app.extractors import extract_coating, extract_din

    text = "Гайка М10 DIN934 A2"
    assert extract_coating(text) == "нержавейка"
    assert extract_din(text) == "DIN 934"


def test_extract_standards_spacing():
    from app.extractors import extract_din

    text = "Винт M6x20 DIN933"
    assert extract_din(text) == "DIN 933"


def test_extract_item_types():
    from app.extractors import extract_item_type

    assert extract_item_type("Болт М12") == "болт"
    assert extract_item_type("Гайка М16") == "гайка"
    assert extract_item_type("Шайба пружинная") == "шайба"
    assert extract_item_type("Шпилька М10") == "шпилька"
    assert extract_item_type("Анкер забивной") == "анкер"
    assert extract_item_type("Канцтовары") == ""


def test_extract_coating_variants():
    from app.extractors import extract_coating

    assert extract_coating("Болт М12 оц") == "цинк"
    assert extract_coating("Болт М12 нерж") == "нержавейка"
    assert extract_coating("Болт М12 латунь") == "латунь"
    assert extract_coating("Болт М12 фосфат") == "фосфат"
    assert extract_coating("Болт М12 черн") == "оксид"
    assert extract_coating("Канцтовары") == ""


def test_extract_strength_space_separated():
    from app.extractors import extract_strength

    assert extract_strength("Болт М12 кл.пр. 8 8") == "8.8"


# ── 8b. Confidence & status ───────────────────────────


def test_confidence_full():
    from app.extractors import compute_confidence, compute_status

    text = "Болт М12-6gx150 8.8 оц ГОСТ 7798-70"
    assert compute_confidence(text) == 5
    assert compute_status(5) == "ok"


def test_confidence_warning():
    from app.extractors import compute_confidence, compute_status

    text = "Болт М12x80"
    conf = compute_confidence(text)
    assert conf == 2  # diameter + length
    assert compute_status(conf) == "review"


def test_confidence_error():
    from app.extractors import compute_confidence, compute_status

    text = "Канцтовары прочие"
    conf = compute_confidence(text)
    assert conf == 0
    assert compute_status(conf) == "manual"


# ── 8c. Transform adds confidence/status columns ─────


def test_transform_has_confidence_columns(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Канцтовары прочие", "Заказ": 5},
    ]
    resp = _upload_file(client, rows)
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["diameter", "strength"]},
    )
    assert resp2.status_code == 200
    html = resp2.text
    # Stats block present with Russian labels
    assert "Не требует проверки:" in html
    assert "Требуется просмотреть:" in html
    assert "Требуется вручную разобрать:" in html
    # Row highlighting attributes
    assert 'data-status="ok"' in html or 'data-status="review"' in html or 'data-status="manual"' in html
    # Filter checkboxes
    assert 'data-filter="ok"' in html


# ── 8c2. No English status words in UI ────────────────


def test_ui_has_no_english_status_words(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Канцтовары прочие", "Заказ": 5},
    ]
    resp = _upload_file(client, rows)
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["diameter", "strength"]},
    )
    assert resp2.status_code == 200
    html = resp2.text

    # No English status words anywhere in rendered HTML
    # (case-sensitive: avoid false positives from CSS class names / data attributes)
    for forbidden in ["OK:", "OK<", ">OK<", "Warning", "Error"]:
        assert forbidden not in html, f"Found forbidden English status text: {forbidden!r}"

    # Russian labels ARE present
    assert "Не требует проверки" in html
    assert "Требуется просмотреть" in html
    assert "Требуется вручную разобрать" in html


# ── 8d. Download OK-only xlsx ─────────────────────────


def test_download_ok_only(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Канцтовары", "Заказ": 5},
    ]
    resp = _upload_file(client, rows)
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["diameter", "strength", "coating", "gost"]},
    )
    assert resp2.status_code == 200
    html = resp2.text

    # Find the OK-only download link
    ok_link = re.search(r'href="(/download/[^"]+)"[^>]*>Скачать только строки без проверки', html)
    assert ok_link, "OK-only download link not found"

    resp3 = client.get(ok_link.group(1))
    assert resp3.status_code == 200
    assert "spreadsheetml" in resp3.headers["content-type"]

    df = pd.read_excel(io.BytesIO(resp3.content), engine="openpyxl")
    # Only OK rows should be present (row 1 with full metiz description)
    # Row 2 ("Канцтовары") has confidence 0 → manual, should be excluded
    for _, row in df.iterrows():
        assert row.get("Статус", "") != "manual"


# ── 9. Download xlsx after transform ────────────────────


def test_download_xlsx_after_transform(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Гайка М16 10.9 DIN 934", "Заказ": 20},
    ]
    resp = _upload_file(client, rows)
    assert resp.status_code == 200
    file_id = _extract_file_id(resp.text)

    # Transform
    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["size", "strength", "coating", "gost"]},
    )
    assert resp2.status_code == 200
    html = resp2.text

    # Extract download link
    m = re.search(r'href="(/download/[^"]+)"', html)
    assert m, "download link not found in result page"
    download_url = m.group(1)

    # Download
    resp3 = client.get(download_url)
    assert resp3.status_code == 200
    assert "spreadsheetml" in resp3.headers["content-type"]

    # Verify it's a valid xlsx with expected columns
    df = pd.read_excel(io.BytesIO(resp3.content), engine="openpyxl")
    assert "Размер MxL" in df.columns
    assert "Класс прочности" in df.columns
    assert "Покрытие" in df.columns
    assert len(df) == 2  # all rows, not limited to 200


# ── 9b. Multi-source signal merging ──────────────────────


def _make_xlsx_with_extra_cols(rows):
    """Create xlsx with extra columns (Стандарт, Класс прочности, Примечание)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = list(rows[0].keys())
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[h])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _upload_extra_cols(client, rows, filename="extra.xlsx"):
    buf = _make_xlsx_with_extra_cols(rows)
    return client.post(
        "/upload",
        files={"file": (filename, buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_transform_uses_strength_column(client):
    """Strength from dedicated column when name lacks it."""
    rows = [
        {"Код": "001", "Номенклатура": "Болт М12x80 оц ГОСТ 7798-70", "Заказ": 10, "Класс прочности": "8.8"},
        {"Код": "002", "Номенклатура": "Гайка М16 DIN 934", "Заказ": 20, "Класс прочности": "10.9"},
    ]
    resp = _upload_extra_cols(client, rows)
    assert resp.status_code == 200
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["diameter", "strength"]},
    )
    assert resp2.status_code == 200
    html = resp2.text
    assert "8.8" in html
    assert "10.9" in html


def test_transform_uses_standard_column(client):
    """Standard from dedicated column: ГОСТ Р ИСО 4014-2013."""
    rows = [
        {"Код": "001", "Номенклатура": "Болт М12x80", "Заказ": 10, "Стандарт": "ГОСТ Р ИСО 4014-2013"},
    ]
    resp = _upload_extra_cols(client, rows)
    assert resp.status_code == 200
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["gost"]},
    )
    assert resp2.status_code == 200
    html = resp2.text
    assert "ГОСТ Р ИСО 4014-2013" in html


def test_transform_standard_number_only(client):
    """Bare number in standard column → treated as ГОСТ."""
    rows = [
        {"Код": "001", "Номенклатура": "Шайба М12", "Заказ": 10, "Стандарт": "11371-78"},
    ]
    resp = _upload_extra_cols(client, rows)
    assert resp.status_code == 200
    file_id = _extract_file_id(resp.text)

    resp2 = client.post(
        "/transform",
        data={"file_id": file_id, "fields": ["gost"]},
    )
    assert resp2.status_code == 200
    html = resp2.text
    assert "ГОСТ 11371-78" in html


def test_preview_shows_extra_columns_if_present(client):
    """If parser detected standard/strength/note columns, they appear in preview."""
    rows = [
        {"Код": "001", "Номенклатура": "Болт М12", "Заказ": 10,
         "Стандарт": "ГОСТ 7798-70", "Класс прочности": "8.8", "Примечание": "срочно"},
    ]
    resp = _upload_extra_cols(client, rows)
    assert resp.status_code == 200
    html = resp.text
    # Preview table should show the raw columns
    assert "ГОСТ 7798-70" in html or "standard_raw" in html
    assert "8.8" in html or "strength_raw" in html


# ── 10. Fallback: manual column selection flow ───────────


def test_fallback_manual_selection(client):
    """Upload file with unrecognizable headers → column selection page → apply → transform."""
    from openpyxl import Workbook

    # Create xlsx with unrecognizable headers
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Alpha")
    ws.cell(row=1, column=2, value="Beta")
    ws.cell(row=1, column=3, value="Gamma")
    ws.cell(row=2, column=1, value="X01")
    ws.cell(row=2, column=2, value="Болт M12x80 8.8")
    ws.cell(row=2, column=3, value=50)
    ws.cell(row=3, column=1, value="X02")
    ws.cell(row=3, column=2, value="Гайка М16 10.9")
    ws.cell(row=3, column=3, value=100)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Step 1: Upload → should get column selection page
    resp = client.post(
        "/upload",
        files={"file": ("bad_headers.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    html = resp.text
    assert "select" in html.lower() or "name_col" in html
    assert "apply-columns" in html

    file_id = _extract_file_id(html)

    # Step 2: Apply columns manually (name=col1, qty=col2, code=col0, header=row0)
    resp2 = client.post(
        "/apply-columns",
        data={"file_id": file_id, "name_col": 1, "qty_col": 2, "code_col": 0, "header_row": 0},
    )
    assert resp2.status_code == 200
    html2 = resp2.text
    # Should now be on view_raw page with checkboxes
    assert 'name="fields"' in html2
    assert "Болт M12x80 8.8" in html2

    file_id2 = _extract_file_id(html2)

    # Step 3: Transform works
    resp3 = client.post(
        "/transform",
        data={"file_id": file_id2, "fields": ["diameter", "strength"]},
    )
    assert resp3.status_code == 200
    html3 = resp3.text
    assert "Диаметр" in html3
    assert "M12" in html3
    assert "8.8" in html3
