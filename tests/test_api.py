import io

import pandas as pd
import pytest
from fastapi.testclient import TestClient


@pytest.fixture(autouse=True)
def _set_dirs(tmp_path, monkeypatch):
    upload_dir = tmp_path / "uploads"
    cache_dir = tmp_path / "cache"
    monkeypatch.setenv("OTDELZAKUP_UPLOAD_DIR", str(upload_dir))
    monkeypatch.setenv("OTDELZAKUP_CACHE_DIR", str(cache_dir))
    import app.cache as cache_mod

    cache_mod.UPLOAD_DIR = upload_dir
    cache_mod.CACHE_DIR = cache_dir


@pytest.fixture()
def client():
    from app.main import app

    return TestClient(app)


def _make_xlsx(rows: list[dict]) -> io.BytesIO:
    buf = io.BytesIO()
    pd.DataFrame(rows).to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf


def _api_upload(client, rows, filename="test.xlsx"):
    xlsx = _make_xlsx(rows)
    return client.post(
        "/api/v1/upload",
        files={"file": (filename, xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


# ── POST /api/v1/upload ─────────────────────────────────────


def test_api_upload_ok(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 100},
        {"Код": "002", "Номенклатура": "Гайка М16 10.9 DIN 934", "Заказ": 200},
    ]
    resp = _api_upload(client, rows)
    assert resp.status_code == 200

    body = resp.json()
    assert "file_id" in body
    assert len(body["file_id"]) == 16
    assert body["filename"] == "test.xlsx"
    assert body["rows_total"] == 2
    assert "name" in body["columns"]
    assert "qty" in body["columns"]
    assert "code" in body["columns"]
    assert body["needs_column_selection"] is False


def test_api_upload_not_xlsx(client):
    buf = io.BytesIO(b"hello")
    resp = client.post(
        "/api/v1/upload",
        files={"file": ("bad.txt", buf, "text/plain")},
    )
    assert resp.status_code == 400
    assert "error" in resp.json()


def test_api_upload_xls_rejected(client):
    buf = io.BytesIO(b"fake xls content")
    resp = client.post(
        "/api/v1/upload",
        files={"file": ("data.xls", buf, "application/vnd.ms-excel")},
    )
    assert resp.status_code == 400
    body = resp.json()
    assert ".xls" in body["error"]


# ── GET /api/v1/preview/{file_id} ───────────────────────────


def test_api_preview_ok(client):
    rows = [{"Код": str(i), "Номенклатура": f"Товар_{i}", "Заказ": i} for i in range(10)]
    resp = _api_upload(client, rows)
    fid = resp.json()["file_id"]

    resp2 = client.get(f"/api/v1/preview/{fid}?limit=5")
    assert resp2.status_code == 200

    body = resp2.json()
    assert body["file_id"] == fid
    assert body["rows_total"] == 10
    assert body["limit"] == 5
    assert len(body["rows"]) == 5
    assert body["columns"] == ["code", "name", "qty", "uom"]
    # Each row is a list, not a dict
    assert isinstance(body["rows"][0], list)


def test_api_preview_default_limit(client):
    rows = [{"Код": str(i), "Номенклатура": f"Товар_{i}", "Заказ": i} for i in range(250)]
    resp = _api_upload(client, rows)
    fid = resp.json()["file_id"]

    resp2 = client.get(f"/api/v1/preview/{fid}")
    body = resp2.json()
    assert body["rows_total"] == 250
    assert body["limit"] == 200
    assert len(body["rows"]) == 200


# ── POST /api/v1/transform ──────────────────────────────────


def test_api_transform_ok(client):
    rows = [
        {"Код": "001", "Номенклатура": "Болт M12x80 8.8 оц ГОСТ 7798-70", "Заказ": 10},
        {"Код": "002", "Номенклатура": "Гайка М16 10.9 DIN 934", "Заказ": 20},
    ]
    resp = _api_upload(client, rows)
    fid = resp.json()["file_id"]

    resp2 = client.post(
        "/api/v1/transform",
        json={"file_id": fid, "fields": ["diameter", "strength"]},
    )
    assert resp2.status_code == 200

    body = resp2.json()
    assert body["file_id"] == fid
    assert body["rows_total"] == 2
    assert body["fields"] == ["diameter", "strength"]
    assert "Диаметр" in body["columns"]
    assert "Класс прочности" in body["columns"]
    # Check extracted values in rows
    flat = [cell for row in body["rows"] for cell in row]
    assert "M12" in flat
    assert "8.8" in flat
    assert "M16" in flat
    assert "10.9" in flat


def test_api_transform_no_fields(client):
    rows = [{"Код": "001", "Номенклатура": "Болт M12x80", "Заказ": 5}]
    resp = _api_upload(client, rows)
    fid = resp.json()["file_id"]

    resp2 = client.post(
        "/api/v1/transform",
        json={"file_id": fid, "fields": []},
    )
    assert resp2.status_code == 200
    body = resp2.json()
    assert body["fields"] == []
    assert "Диаметр" not in body["columns"]


def test_api_transform_with_limit(client):
    rows = [{"Код": str(i), "Номенклатура": f"Болт M{i}x10", "Заказ": 1} for i in range(50)]
    resp = _api_upload(client, rows)
    fid = resp.json()["file_id"]

    resp2 = client.post(
        "/api/v1/transform",
        json={"file_id": fid, "fields": ["diameter"], "limit": 10},
    )
    assert resp2.status_code == 200
    body = resp2.json()
    assert body["rows_total"] == 50
    assert len(body["rows"]) == 10
    assert "Диаметр" in body["columns"]


# ── 404 for unknown file_id ─────────────────────────────────


def test_api_preview_not_found(client):
    resp = client.get("/api/v1/preview/0000000000000000")
    assert resp.status_code == 404
    assert resp.json()["error"] == "not found"


def test_api_transform_not_found(client):
    resp = client.post(
        "/api/v1/transform",
        json={"file_id": "0000000000000000", "fields": ["diameter"]},
    )
    assert resp.status_code == 404
    assert resp.json()["error"] == "not found"


# ── Determinism: same file → same file_id ────────────────────


def test_api_upload_deterministic_file_id(client):
    rows = [{"Код": "001", "Номенклатура": "Тест", "Заказ": 1}]
    xlsx = _make_xlsx(rows)
    raw = xlsx.read()

    def _upload_same_bytes():
        buf = io.BytesIO(raw)
        return client.post(
            "/api/v1/upload",
            files={"file": ("one.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    resp1 = _upload_same_bytes()
    resp2 = _upload_same_bytes()
    # Same bytes → same file_id
    assert resp1.json()["file_id"] == resp2.json()["file_id"]


# ── API fallback: needs_column_selection ─────────────────────


def _upload_bad_headers_xlsx(client, filename="bad.xlsx"):
    """Upload an xlsx with unrecognizable headers via API."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Alpha")
    ws.cell(row=1, column=2, value="Beta")
    ws.cell(row=1, column=3, value="Gamma")
    ws.cell(row=2, column=1, value="X01")
    ws.cell(row=2, column=2, value="Болт М12x80")
    ws.cell(row=2, column=3, value=50)
    ws.cell(row=3, column=1, value="X02")
    ws.cell(row=3, column=2, value="Гайка М16")
    ws.cell(row=3, column=3, value=100)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return client.post(
        "/api/v1/upload",
        files={"file": (filename, buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_api_upload_needs_column_selection(client):
    """Upload with unrecognizable headers returns needs_column_selection=True."""
    resp = _upload_bad_headers_xlsx(client)
    assert resp.status_code == 200

    body = resp.json()
    assert body["needs_column_selection"] is True
    assert "file_id" in body
    assert "preview_rows" in body
    assert "num_columns" in body
    assert body["num_columns"] >= 3
    assert "detected" in body


def test_api_apply_columns_ok(client):
    """Apply columns after fallback, then verify transform works."""
    resp = _upload_bad_headers_xlsx(client)
    body = resp.json()
    fid = body["file_id"]

    # Apply: name=col1, qty=col2, code=col0, header=row0
    resp2 = client.post(
        "/api/v1/apply-columns",
        json={"file_id": fid, "name_col": 1, "qty_col": 2, "code_col": 0, "header_row": 0},
    )
    assert resp2.status_code == 200
    body2 = resp2.json()
    assert body2["rows_total"] == 2
    assert "name" in body2["columns"]
    assert "code" in body2["columns"]

    # Transform should work now
    resp3 = client.post(
        "/api/v1/transform",
        json={"file_id": fid, "fields": ["diameter"]},
    )
    assert resp3.status_code == 200
    body3 = resp3.json()
    assert "Диаметр" in body3["columns"]
    flat = [cell for row in body3["rows"] for cell in row]
    assert "M12" in flat


def test_api_apply_columns_not_found(client):
    """Apply columns for unknown file_id returns 404."""
    resp = client.post(
        "/api/v1/apply-columns",
        json={"file_id": "0000000000000000", "name_col": 0, "qty_col": 1, "header_row": 0},
    )
    assert resp.status_code == 404
    assert resp.json()["error"] == "not found"
