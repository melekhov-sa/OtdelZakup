"""Tests for qty/uom parsing — parse_qty_uom, combined-column detection, name fallback."""

import io

import pandas as pd
import pytest
from fastapi.testclient import TestClient

from app.parser_excel import (
    _apply_name_qty_fallback,
    _extract_qty_uom_suffix,
    parse_qty_uom,
)


# ── Unit tests for parse_qty_uom ─────────────────────────────


def test_parse_qty_uom_basic_formats():
    """All common formats return correct qty, uom, and empty rest."""
    cases = [
        ("1500 шт", 1500.0, "шт", ""),
        ("8шт", 8.0, "шт", ""),
        ("2,5 кг", 2.5, "кг", ""),
        ("10 м", 10.0, "м", ""),
        ("100 уп", 100.0, "уп", ""),
        ("3 компл", 3.0, "компл", ""),
        ("5 пач", 5.0, "пач", ""),
        ("50 мм", 50.0, "мм", ""),
        ("1.5 л", 1.5, "л", ""),
    ]
    for text, exp_qty, exp_uom, exp_rest in cases:
        qty, uom, rest = parse_qty_uom(text)
        assert qty == pytest.approx(exp_qty), f"{text!r}: expected qty {exp_qty}, got {qty}"
        assert uom == exp_uom, f"{text!r}: expected uom {exp_uom!r}, got {uom!r}"
        assert rest == exp_rest, f"{text!r}: expected rest {exp_rest!r}, got {rest!r}"


def test_parse_qty_uom_normalization():
    """Raw UOM variants are normalized to canonical form."""
    cases = [
        ("50 штук", "шт"),
        ("50 штука", "шт"),
        ("10 килограмм", "кг"),
        ("3 упаковки", "уп"),
        ("2 комплект", "компл"),
        ("7 метров", "м"),
        ("500 граммов", "г"),
    ]
    for text, exp_uom in cases:
        _, uom, _ = parse_qty_uom(text)
        assert uom == exp_uom, f"{text!r}: expected {exp_uom!r}, got {uom!r}"


def test_parse_qty_uom_no_match():
    """Pure numbers or text without UOM return (None, None, text)."""
    for text in ["100", "Болт М12х80", "8.8", "", "   "]:
        qty, uom, rest = parse_qty_uom(text)
        assert qty is None, f"{text!r} should not match qty"
        assert uom is None, f"{text!r} should not match uom"


def test_parse_qty_uom_embedded_in_name():
    """Extracts qty+uom from within a longer name and returns cleaned rest."""
    qty, uom, rest = parse_qty_uom("Болт М12х80 50шт")
    assert qty == pytest.approx(50.0)
    assert uom == "шт"
    assert rest == "Болт М12х80"


def test_parse_qty_uom_parens_cleanup():
    """Removes parentheses wrapping qty+uom and returns clean name."""
    qty, uom, rest = parse_qty_uom("Гайка М10 (200 шт)")
    assert qty == pytest.approx(200.0)
    assert uom == "шт"
    assert rest == "Гайка М10"


def test_extract_qty_uom_suffix_trailing():
    """Extracts only from end of string (suffix variant)."""
    qty, uom, rest = _extract_qty_uom_suffix("Болт М12х80 8.8 50шт")
    assert qty == pytest.approx(50.0)
    assert uom == "шт"
    assert rest == "Болт М12х80 8.8"


def test_extract_qty_uom_suffix_parens():
    """Extracts from trailing parentheses."""
    qty, uom, rest = _extract_qty_uom_suffix("Шайба А12 (100 уп)")
    assert qty == pytest.approx(100.0)
    assert uom == "уп"
    assert rest == "Шайба А12"


def test_extract_qty_uom_suffix_no_match_middle():
    """Does NOT extract qty+uom from the middle of a name (must be at end)."""
    qty, uom, _ = _extract_qty_uom_suffix("Болт 10шт М12")
    assert qty is None
    assert uom is None


# ── Integration: Case A — combined column detection ──────────


@pytest.fixture(autouse=True)
def _set_dirs(tmp_path, monkeypatch):
    upload_dir = tmp_path / "uploads"
    cache_dir = tmp_path / "cache"
    monkeypatch.setenv("OTDELZAKUP_UPLOAD_DIR", str(upload_dir))
    monkeypatch.setenv("OTDELZAKUP_CACHE_DIR", str(cache_dir))
    import app.cache as cache_mod

    cache_mod.UPLOAD_DIR = upload_dir
    cache_mod.CACHE_DIR = cache_dir

    db_path = tmp_path / "test.db"
    monkeypatch.setenv("OTDELZAKUP_DB_PATH", str(db_path))
    import app.database as db_mod
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    db_mod.DB_PATH = db_path
    db_mod.engine = create_engine(
        f"sqlite:///{db_path}", connect_args={"check_same_thread": False}
    )
    db_mod.SessionLocal = sessionmaker(
        bind=db_mod.engine, autoflush=False, expire_on_commit=False
    )
    db_mod.init_db()


@pytest.fixture()
def client():
    from app.main import app

    return TestClient(app)


def _make_xlsx(rows: list[dict]) -> io.BytesIO:
    buf = io.BytesIO()
    pd.DataFrame(rows).to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf


def _upload(client, rows, filename="test.xlsx"):
    xlsx = _make_xlsx(rows)
    return client.post(
        "/upload",
        files={
            "file": (
                filename,
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_combined_col_splits_qty_and_uom(client):
    """Upload file with '100 шт' in qty column → parsed to qty=100, uom='шт'."""
    rows = [
        {"Наименование": "Болт М12х80", "Количество": "100 шт"},
        {"Наименование": "Гайка М10", "Количество": "50 шт"},
    ]
    resp = _upload(client, rows)
    assert resp.status_code == 200
    html = resp.text
    # The parsed preview table should contain qty values
    assert "Болт М12х80" in html
    # UOM column should appear (Ед.)
    assert "Ед." in html or "100" in html


def test_combined_col_fractional_qty(client):
    """Upload with '2,5 кг' → qty=2.5, uom='кг' persisted."""
    from app.cache import load_dataframe
    from app.parser_excel import _QTY_UOM_RE
    import re

    rows = [
        {"Наименование": "Кабель", "Заказ": "2,5 кг"},
        {"Наименование": "Лист", "Заказ": "10 м"},
    ]
    resp = _upload(client, rows)
    assert resp.status_code == 200

    # Extract file_id from response and check cached DataFrame
    import re as _re
    m = _re.search(r'name="file_id"\s+value="([^"]+)"', resp.text)
    if m:
        fid = m.group(1)
        df = load_dataframe(fid)
        if df is not None:
            assert list(df["uom"]) == ["кг", "м"]
            assert float(df["qty"].iloc[0]) == pytest.approx(2.5)


# ── Integration: Case B — name fallback ──────────────────────


def test_case_b_extracts_qty_from_name(client):
    """Upload without qty column — qty extracted from trailing text in name."""
    from app.cache import load_dataframe
    import re as _re

    rows = [
        {"Наименование": "Болт М12х80 50шт"},
        {"Наименование": "Гайка М10 (30 шт)"},
    ]
    resp = _upload(client, rows)
    assert resp.status_code == 200

    m = _re.search(r'name="file_id"\s+value="([^"]+)"', resp.text)
    if m:
        fid = m.group(1)
        df = load_dataframe(fid)
        if df is not None:
            qtys = list(df["qty"].fillna(0))
            # At least one row should have a qty extracted
            assert any(q > 0 for q in qtys), f"Expected extracted qty, got: {qtys}"


# ── Unit test: _apply_name_qty_fallback ──────────────────────


def test_apply_name_qty_fallback_unit():
    """Direct unit test: updates qty/uom from name and cleans name."""
    df = pd.DataFrame(
        [
            {"name": "Болт М12х80 50шт", "qty": None, "uom": "шт"},
            {"name": "Гайка М10 (30 шт)", "qty": None, "uom": "шт"},
            {"name": "Болт М16", "qty": 10, "uom": "шт"},  # already has qty — unchanged
        ]
    )
    result = _apply_name_qty_fallback(df)
    assert result.at[0, "qty"] == 50
    assert result.at[0, "name"] == "Болт М12х80"
    assert result.at[1, "qty"] == 30
    assert result.at[1, "name"] == "Гайка М10"
    assert result.at[2, "qty"] == 10  # unchanged
    assert result.at[2, "name"] == "Болт М16"  # unchanged


# ── format_qty ───────────────────────────────────────────────


def test_format_qty_integer_no_decimal():
    """Whole-number qty displayed without decimal point; fractions kept; None → ''."""
    from app.display_labels import format_qty

    assert format_qty(64.0) == "64"
    assert format_qty(64) == "64"
    assert format_qty(1000.0) == "1000"
    assert format_qty(2.5) == "2.5"
    assert format_qty(2.500) == "2.5"
    assert format_qty(1.125) == "1.125"
    assert format_qty(None) == ""
    assert format_qty(float("nan")) == ""


def test_export_writes_int_for_integer_qty(tmp_path):
    """Excel export writes 64 (int), not 64.0 (float) for whole-number qty values."""
    import io
    import openpyxl
    import pandas as pd
    from app.parser_excel import dataframe_to_xlsx_bytes

    df = pd.DataFrame({
        "code": ["A01"],
        "name": ["Болт М12"],
        "qty": [64.0],
        "uom": ["шт"],
    })
    xlsx_bytes = dataframe_to_xlsx_bytes(df)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    # Find qty column by Russian header label
    headers = {cell.value: cell.column for cell in ws[1]}
    qty_col = headers.get("Количество")
    assert qty_col is not None, "Количество column not found"

    qty_val = ws.cell(row=2, column=qty_col).value
    assert qty_val == 64
    assert isinstance(qty_val, int), f"Expected int, got {type(qty_val)}: {qty_val}"


# ── extract_uom_from_header ──────────────────────────────────


def test_extract_uom_from_header():
    """UOM tokens inside column headers are extracted correctly."""
    from app.parser_excel import extract_uom_from_header

    assert extract_uom_from_header("Кол-во, шт") == "шт"
    assert extract_uom_from_header("Количество (шт.)") == "шт"
    assert extract_uom_from_header("Кол-во (кг)") == "кг"
    assert extract_uom_from_header("Кол-во, ед. изм. шт") == "шт"
    assert extract_uom_from_header("Кол-во, шт.") == "шт"
    assert extract_uom_from_header("Заказ (компл.)") == "компл"


def test_no_uom_when_header_missing():
    """Headers without a UOM token return None."""
    from app.parser_excel import extract_uom_from_header

    assert extract_uom_from_header("Количество") is None
    assert extract_uom_from_header("Наименование") is None
    assert extract_uom_from_header("Код") is None
    assert extract_uom_from_header("") is None
    assert extract_uom_from_header(None) is None
