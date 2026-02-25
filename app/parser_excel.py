import io
import math
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz as _fuzz
except ImportError:  # pragma: no cover – graceful degradation
    _fuzz = None


class ParseError(Exception):
    """Raised when the Excel file structure cannot be parsed."""


# ── Result data types ────────────────────────────────────────


@dataclass
class DetectedColumns:
    """Which columns were detected and how."""
    name_idx: Optional[int] = None
    qty_idx: Optional[int] = None
    uom_idx: Optional[int] = None   # separate unit-of-measure column (e.g. "Ед. изм.")
    code_idx: Optional[int] = None
    standard_idx: Optional[int] = None
    strength_col_idx: Optional[int] = None
    note_idx: Optional[int] = None
    score: int = 0
    header_row: Optional[int] = None
    method: str = "auto"  # "auto", "fuzzy", "heuristic", "manual"
    fuzzy_scores: dict = field(default_factory=dict)
    qty_uom_combined: bool = False  # True when qty column contains "N uom" values
    low_confidence: bool = False    # True when detection is uncertain
    reasons: dict = field(default_factory=dict)  # UI hints per column role


@dataclass
class ParseResult:
    """Result of parsing an Excel file."""
    df: Optional[pd.DataFrame] = None
    raw_values: Optional[list] = None
    raw_headers: Optional[list] = None
    detected: DetectedColumns = field(default_factory=DetectedColumns)
    needs_manual_selection: bool = False


# ── Synonym lists (lowercase) ────────────────────────────────

_CODE_SYNONYMS = ["код", "артикул", "№", "номер"]
_NAME_SYNONYMS = ["номенклатура", "наименование", "товар", "позиция"]
# Note: "зaказ" with latin 'a' included intentionally for OCR/copy-paste issues
_QTY_SYNONYMS = [
    "заказ", "зaказ", "кол", "кол-во", "количество",
    "кол во", "колво", "потребность", "к заказу",
]

_STANDARD_SYNONYMS = ["стандарт", "гост", "iso", "дин", "din"]
_STRENGTH_COL_SYNONYMS = ["класс прочности", "класс", "прочность"]
_NOTE_SYNONYMS = ["примечание", "комментарий", "требование", "тд", "условие"]

_SYNONYMS_MAP = {
    "code": _CODE_SYNONYMS,
    "name": _NAME_SYNONYMS,
    "qty": _QTY_SYNONYMS,
}

# Sub-header tokens that indicate a multiline header
_SUBHEADER_TOKENS = ["шт", "ед.изм", "ед. изм", "единица", "штук"]


# ── Quantity + Unit-of-Measure parsing ────────────────────────

_UOM_NORMALIZED: dict[str, list[str]] = {
    "шт":    ["штук", "штука", "штуки", "шт"],
    "кг":    ["кг", "килограмм", "килограмма", "килограммов"],
    "г":     ["граммов", "грамма", "грамм", "гр", "г"],
    "м":     ["метров", "метра", "метр", "м"],
    "мм":    ["мм"],
    "л":     ["литров", "литра", "литр", "л"],
    "уп":    ["упаковок", "упаковки", "упаковка", "упак", "уп"],
    "компл": ["комплектов", "комплекта", "комплект", "компл"],
    "пач":   ["пачек", "пачки", "пачка", "пач"],
    "м²":    ["кв.м", "м²", "м2"],
    "м³":    ["куб.м", "м³", "м3"],
}

# Flat map: raw_form_lower -> normalized_UOM
_UOM_MAP: dict[str, str] = {}
for _uom_n, _uom_raws in _UOM_NORMALIZED.items():
    for _uom_r in _uom_raws:
        _UOM_MAP[_uom_r.lower()] = _uom_n

# Alternatives sorted longest-first to avoid partial matches (e.g. "мм" before "м")
_UOM_ALTS = sorted(_UOM_MAP.keys(), key=len, reverse=True)

# Synonyms used to detect a dedicated unit-of-measure column header
_UOM_SYNONYMS = ["ед. изм", "ед.изм", "ед.", "ед ", "единица", "единицы"]


def normalize_uom(text: str) -> Optional[str]:
    """Normalize a raw unit-of-measure string to a canonical form.

    Returns the canonical UOM string, or None if the text is not recognized.
    Strips trailing dots and lowercases before lookup.
    """
    if not text:
        return None
    clean = text.strip().lower().rstrip(".")
    return _UOM_MAP.get(clean)


def extract_uom_from_header(header_text: str) -> Optional[str]:
    """Extract unit-of-measure embedded in a column header string.

    Recognizes patterns like:
      "Кол-во, шт"         → "шт"
      "Количество (шт.)"   → "шт"
      "Кол-во (кг)"        → "кг"
      "Кол-во, ед. изм. шт"→ "шт"
      "Кол-во, шт."        → "шт"

    Returns normalized UOM or None if no UOM token is found.
    """
    if not header_text:
        return None
    h = header_text.strip()

    # Strategy 1: the last whitespace/punctuation-separated token
    tokens = re.split(r"[\s,().;]+", h)
    for token in reversed(tokens):
        t = token.strip().lower()
        if t and t in _UOM_MAP:
            return _UOM_MAP[t]

    # Strategy 2: scan tokens after the first separator (comma/paren)
    for sep in (",", "("):
        idx = h.find(sep)
        if idx >= 0:
            remainder = h[idx + 1:].rstrip(")")
            for token in re.split(r"[\s.,();]+", remainder):
                t = token.strip().lower()
                if t and t in _UOM_MAP:
                    return _UOM_MAP[t]

    return None


# General qty+uom pattern: <number> <uom> followed by space, end, or punctuation
_QTY_UOM_RE = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*("
    + "|".join(re.escape(a) for a in _UOM_ALTS)
    + r")\.?(?=\s|$|[,;)])",
    re.IGNORECASE | re.UNICODE,
)

# Suffix pattern: qty+uom in trailing parens or at end of text
_QTY_UOM_SUFFIX_RE = re.compile(
    r"(?:"
    r"\(\s*(\d+(?:[.,]\d+)?)\s*("
    + "|".join(re.escape(a) for a in _UOM_ALTS)
    + r")\.?\s*\)"
    r"|\s+(\d+(?:[.,]\d+)?)\s*("
    + "|".join(re.escape(a) for a in _UOM_ALTS)
    + r")\.?"
    r")\s*$",
    re.IGNORECASE | re.UNICODE,
)


def parse_qty_uom(text: str) -> tuple[Optional[float], Optional[str], str]:
    """Parse quantity and unit of measure from text.

    Returns (qty, uom_normalized, rest_text).
    rest_text is the original text with the matched portion removed.
    If no match: (None, None, text).
    """
    if not text or not text.strip():
        return None, None, text

    m = _QTY_UOM_RE.search(text)
    if not m:
        return None, None, text

    qty_str = m.group(1).replace(",", ".")
    try:
        qty_f = float(qty_str)
    except ValueError:
        return None, None, text

    uom_raw = m.group(2).lower()
    uom_norm = _UOM_MAP.get(uom_raw, uom_raw)

    start, end = m.start(), m.end()
    before = text[:start].rstrip()
    after = text[end:].lstrip()

    # Clean up empty parentheses left behind after removing "200 шт" from "(200 шт)"
    if before.endswith("(") and after.startswith(")"):
        before = before[:-1].rstrip()
        after = after[1:].lstrip()

    rest = (before + " " + after).strip() if (before and after) else (before or after)
    return qty_f, uom_norm, rest


def extract_qty_uom_suffix(text: str) -> tuple[Optional[float], Optional[str], str]:
    """Extract qty+uom specifically from trailing parentheses or end of text.

    More conservative than parse_qty_uom — used for name-column fallback.
    Returns (qty, uom_normalized, cleaned_text).
    """
    m = _QTY_UOM_SUFFIX_RE.search(text)
    if not m:
        return None, None, text

    # Two alternations: parens (groups 1,2) vs space-suffix (groups 3,4)
    qty_str = m.group(1) or m.group(3)
    uom_raw = (m.group(2) or m.group(4) or "").lower()

    if qty_str is None:
        return None, None, text

    try:
        qty_f = float(qty_str.replace(",", "."))
    except ValueError:
        return None, None, text

    uom_norm = _UOM_MAP.get(uom_raw, uom_raw) if uom_raw else None
    rest = text[:m.start()].rstrip()
    return qty_f, uom_norm, rest


# ── Low-level helpers ─────────────────────────────────────────

def _cell_value(ws, row: int, col: int):
    """Get cell value, resolving merged cells to the top-left value."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def _read_sheet_values(ws, max_rows: int | None = None) -> list[list]:
    """Read worksheet into a 2D list, resolving merged cells."""
    last_row = ws.max_row or 0
    if max_rows is not None:
        last_row = min(last_row, max_rows)
    max_col = ws.max_column or 0
    rows = []
    for r in range(1, last_row + 1):
        row = [_cell_value(ws, r, c) for c in range(1, max_col + 1)]
        rows.append(row)
    return rows


def _norm(val) -> str:
    """Normalize a cell value to lowercase stripped string with NFKC."""
    if val is None:
        return ""
    s = str(val).strip().lower()
    return unicodedata.normalize("NFKC", s)


def _matches(header: str, synonyms: list[str]) -> bool:
    """Check if a normalized header contains any synonym."""
    for syn in synonyms:
        if syn in header:
            return True
    return False


# ── Fuzzy matching ────────────────────────────────────────────

def _fuzzy_match(header: str, synonyms: list[str], threshold: int = 70) -> tuple[bool, int]:
    """Check fuzzy similarity of header against synonyms.

    Returns (matched, best_score). Falls back to substring if rapidfuzz unavailable.
    """
    if not header:
        return False, 0

    # Fast path: exact substring match
    for syn in synonyms:
        if syn in header:
            return True, 100

    if _fuzz is None:
        return False, 0

    best = 0
    for syn in synonyms:
        score = _fuzz.partial_ratio(header, syn)
        if score > best:
            best = score
    return best >= threshold, int(best)


def _fuzzy_find_col(headers: list[str], synonyms: list[str], threshold: int = 70) -> tuple[int | None, int]:
    """Find the column with best fuzzy match above threshold.

    Returns (col_index, best_score).
    """
    best_idx = None
    best_score = 0
    for i, h in enumerate(headers):
        if not h:
            continue
        matched, score = _fuzzy_match(h, synonyms, threshold)
        if matched and score > best_score:
            best_score = score
            best_idx = i
    return best_idx, best_score


# ── Header / column detection ─────────────────────────────────

def _table_score(row_values: list) -> int:
    """Calculate table score for a row: +1 per matched column category."""
    headers = [_norm(v) for v in row_values]
    score = 0
    for _category, synonyms in _SYNONYMS_MAP.items():
        for h in headers:
            if h:
                matched, _ = _fuzzy_match(h, synonyms, threshold=70)
                if matched:
                    score += 1
                    break
    return score


def _find_header_row_scored(values_2d: list[list], max_scan: int = 80) -> tuple[int | None, int]:
    """Find the best header row by table score.

    Returns (row_index, score). Only considers rows with >=2 non-empty cells.
    """
    best_idx = None
    best_score = 0
    for i, row in enumerate(values_2d[:max_scan]):
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        if non_empty < 2:
            continue
        score = _table_score(row)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx, best_score


def _check_multiline_header(values_2d: list[list], header_idx: int) -> tuple[list[str], bool]:
    """Check for multiline header and concatenate if needed.

    Returns (final_headers, consumed_next_row).
    """
    headers = [_norm(v) for v in values_2d[header_idx]]
    next_idx = header_idx + 1

    if next_idx >= len(values_2d):
        return headers, False

    next_row = [_norm(v) for v in values_2d[next_idx]]

    # Check if the next row contains sub-header tokens
    has_subheader = False
    for cell in next_row:
        if not cell:
            continue
        # Skip cells that look like qty+uom data values (e.g. "100 шт", "2,5 кг")
        # to avoid mistaking data rows for multiline header continuation
        _, _uom_check, _ = parse_qty_uom(cell)
        if _uom_check is not None:
            continue
        for token in _SUBHEADER_TOKENS:
            if token in cell:
                has_subheader = True
                break
        # Also check if it looks like a continuation (has qty-like words not in main header)
        if not has_subheader:
            for syn in _QTY_SYNONYMS:
                if syn in cell and not any(syn in h for h in headers if h):
                    has_subheader = True
                    break
        if has_subheader:
            break

    if not has_subheader:
        return headers, False

    # Concatenate: header[col] + " " + next_row[col]
    merged = []
    for i in range(len(headers)):
        parts = []
        if headers[i]:
            parts.append(headers[i])
        if i < len(next_row) and next_row[i]:
            parts.append(next_row[i])
        merged.append(" ".join(parts))
    return merged, True


def _find_col(headers: list[str], synonyms: list[str]) -> int | None:
    """Find column by substring match (legacy fast path)."""
    for i, h in enumerate(headers):
        if h and _matches(h, synonyms):
            return i
    return None


# ── Content heuristics ────────────────────────────────────────

def _find_qty_heuristic(
    values_2d: list[list], data_start: int, num_cols: int, exclude_cols: set[int] | None = None
) -> int | None:
    """Fallback: find a column where >=60% of values are numeric and mean > 0."""
    exclude = exclude_cols or set()
    total = len(values_2d) - data_start
    if total <= 0:
        return None
    best_idx, best_count = None, 0
    for col_idx in range(num_cols):
        if col_idx in exclude:
            continue
        count = 0
        value_sum = 0.0
        for row_idx in range(data_start, len(values_2d)):
            val = values_2d[row_idx][col_idx] if col_idx < len(values_2d[row_idx]) else None
            if val is not None:
                try:
                    num = float(val)
                    count += 1
                    value_sum += num
                except (ValueError, TypeError):
                    pass
        if count > best_count and count >= total * 0.6:
            mean = value_sum / count if count > 0 else 0
            if mean > 0:
                best_count = count
                best_idx = col_idx
    return best_idx


def _find_name_heuristic(
    values_2d: list[list], data_start: int, num_cols: int, exclude_cols: set[int] | None = None
) -> int | None:
    """Fallback: find column with longest average string length that is mostly text."""
    exclude = exclude_cols or set()
    total = len(values_2d) - data_start
    if total <= 0:
        return None
    best_idx = None
    best_avg_len = 0.0
    for col_idx in range(num_cols):
        if col_idx in exclude:
            continue
        text_count = 0
        total_len = 0
        for row_idx in range(data_start, len(values_2d)):
            val = values_2d[row_idx][col_idx] if col_idx < len(values_2d[row_idx]) else None
            if val is not None:
                s = str(val).strip()
                if s:
                    # Check if it's not purely numeric
                    try:
                        float(s)
                    except (ValueError, TypeError):
                        text_count += 1
                        total_len += len(s)
        if text_count > total * 0.5 and text_count > 0:
            avg_len = total_len / text_count
            if avg_len > best_avg_len:
                best_avg_len = avg_len
                best_idx = col_idx
    return best_idx


# ── Combined qty+uom column helpers ──────────────────────────

def _combined_ratio(values_2d: list[list], data_start: int, col_idx: int) -> float:
    """Return fraction of non-empty cells in col_idx that match a qty+uom pattern."""
    total_nonnull = 0
    combined_count = 0
    for row_idx in range(data_start, len(values_2d)):
        val = values_2d[row_idx][col_idx] if col_idx < len(values_2d[row_idx]) else None
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue
        total_nonnull += 1
        _, uom, _ = parse_qty_uom(s)
        if uom is not None:
            combined_count += 1
    return combined_count / total_nonnull if total_nonnull > 0 else 0.0


def _find_qty_uom_combined_heuristic(
    values_2d: list[list], data_start: int, num_cols: int, exclude_cols: set[int] | None = None
) -> int | None:
    """Find a column where >=50% of non-empty values look like 'N uom'."""
    exclude = exclude_cols or set()
    best_idx, best_ratio = None, 0.0
    for col_idx in range(num_cols):
        if col_idx in exclude:
            continue
        ratio = _combined_ratio(values_2d, data_start, col_idx)
        if ratio >= 0.5 and ratio > best_ratio:
            best_ratio = ratio
            best_idx = col_idx
    return best_idx


# Backward-compatible private alias
_extract_qty_uom_suffix = extract_qty_uom_suffix


def _apply_name_qty_fallback(df: pd.DataFrame) -> pd.DataFrame:
    """For rows with no qty, try to extract qty+uom from trailing portion of name."""
    mask = df["qty"].isna()
    if not mask.any():
        return df
    df = df.copy()
    for idx in df[mask].index:
        name = str(df.at[idx, "name"])
        qty, uom, rest = extract_qty_uom_suffix(name)
        if qty is not None:
            df.at[idx, "qty"] = qty if qty != int(qty) else int(qty)
            if uom:
                df.at[idx, "uom"] = uom
            if rest.strip():
                df.at[idx, "name"] = rest.strip()
    return df


# ── DataFrame builder ─────────────────────────────────────────

def build_dataframe_from_columns(
    values_2d: list[list],
    header_idx: int,
    name_idx: int,
    qty_idx: int | None,
    code_idx: int | None,
    standard_idx: int | None = None,
    strength_col_idx: int | None = None,
    note_idx: int | None = None,
    qty_is_combined: bool = False,
    uom_idx: int | None = None,
) -> pd.DataFrame:
    """Build canonical DataFrame from raw 2D values and explicit column indices.

    Skips rows where name is empty. Returns DataFrame with columns:
    code, name, qty, uom, standard_raw, strength_raw, note_raw,
    raw_text, qty_uom_source.
    Raises ParseError if no data rows found.

    Uses RowParser for per-row parsing (strict qty/uom: both or none, no defaults).
    """
    from app.parsing.row_parser import parse_row  # noqa: PLC0415 – avoid circular at module level

    # Extract header row as a list of strings
    header_row = values_2d[header_idx] if header_idx < len(values_2d) else []

    def _safe_hdr(idx: int | None) -> str | None:
        """Return the header string for column index (or None if idx is None)."""
        if idx is None:
            return None
        h = header_row[idx] if idx < len(header_row) else None
        if h is None or (isinstance(h, float) and math.isnan(h)):
            return f"_col_{idx}"
        s = str(h).strip()
        return s if s else f"_col_{idx}"

    # Use sentinel keys for mapped columns to guarantee uniqueness even if headers clash
    name_key = f"__name__{name_idx}"
    qty_key = f"__qty__{qty_idx}" if qty_idx is not None else None
    uom_key = f"__uom__{uom_idx}" if uom_idx is not None else None
    code_key = f"__code__{code_idx}" if code_idx is not None else None
    std_key = f"__std__{standard_idx}" if standard_idx is not None else None
    str_key = f"__str__{strength_col_idx}" if strength_col_idx is not None else None
    note_key = f"__note__{note_idx}" if note_idx is not None else None

    # Extract UOM embedded in the qty column header (e.g. "Кол-во, шт" → "шт")
    # Only used when there is no dedicated uom_col and the column is not combined.
    qty_header_uom: Optional[str] = None
    if qty_idx is not None and uom_idx is None and not qty_is_combined:
        raw_qty_hdr = header_row[qty_idx] if qty_idx < len(header_row) else None
        if raw_qty_hdr is not None:
            qty_header_uom = extract_uom_from_header(str(raw_qty_hdr))

    mapped_indices = frozenset(
        i for i in (name_idx, qty_idx, uom_idx, code_idx, standard_idx, strength_col_idx, note_idx)
        if i is not None
    )

    mapping = {
        "name_col": name_key,
        "qty_col": qty_key,
        "uom_col": uom_key,
        "code_col": code_key,
        "standard_col": std_key,
        "strength_col": str_key,
        "note_col": note_key,
        "qty_is_combined": qty_is_combined,
        "qty_header_uom": qty_header_uom,
    }

    data_start = header_idx + 1
    result_rows = []

    for row in values_2d[data_start:]:
        def _get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        # Sentinel-keyed cells for mapped columns
        cells: dict = {
            name_key: _get(name_idx),
        }
        if qty_key:
            cells[qty_key] = _get(qty_idx)
        if uom_key:
            cells[uom_key] = _get(uom_idx)
        if code_key:
            cells[code_key] = _get(code_idx)
        if std_key:
            cells[std_key] = _get(standard_idx)
        if str_key:
            cells[str_key] = _get(strength_col_idx)
        if note_key:
            cells[note_key] = _get(note_idx)

        # Unmapped columns contribute to raw_text
        for col_idx, val in enumerate(row):
            if col_idx in mapped_indices:
                continue
            vs = str(val).strip() if val is not None else ""
            if not vs or (isinstance(val, float) and math.isnan(val)):
                continue
            cells[f"_extra_{col_idx}"] = val

        parsed = parse_row(cells, mapping)

        if not parsed["name_raw"]:
            continue

        result_rows.append({
            "code": parsed["code"] or "",
            "name": parsed["name"],
            "qty": parsed["qty"],
            "uom": parsed["uom"],
            "standard_raw": parsed["standard_raw"] or "",
            "strength_raw": parsed["strength_raw"] or "",
            "note_raw": parsed["note_raw"] or "",
            "raw_text": parsed["raw_text"],
            "qty_uom_source": parsed["qty_uom_source"],
        })

    if not result_rows:
        raise ParseError("Табличная часть найдена, но данные отсутствуют.")

    return pd.DataFrame(result_rows)


# ── Main parser (new entry point) ─────────────────────────────

_PARSE_ERR = (
    "Не найдена табличная часть "
    "(ожидаются колонки: Код, Номенклатура, Кол-во/Заказ)."
)


def parse_excel(file_path: str | Path) -> ParseResult:
    """Parse .xlsx with smart auto-detect. Returns ParseResult (never raises on detection failure)."""
    wb = load_workbook(str(file_path), data_only=True)
    ws = wb.active
    values_2d = _read_sheet_values(ws)
    wb.close()

    if not values_2d:
        raise ParseError("Файл пуст.")

    num_cols = max(len(row) for row in values_2d) if values_2d else 0

    # Step 1: find best header row by score
    header_idx, score = _find_header_row_scored(values_2d)

    detected = DetectedColumns(score=score, header_row=header_idx)

    if header_idx is None or score < 2:
        # Cannot find a confident header row → fallback
        # Use content-based scorer to pre-fill column guesses
        from app.column_scorer import run_column_scorer  # noqa: PLC0415
        sr = run_column_scorer(values_2d)
        detected.name_idx = sr.name_idx
        detected.qty_idx = sr.qty_idx
        detected.uom_idx = sr.uom_idx
        detected.code_idx = sr.code_idx
        if sr.header_row is not None:
            detected.header_row = sr.header_row
        detected.low_confidence = True
        detected.reasons = sr.reasons
        raw_headers = [str(v) if v is not None else "" for v in values_2d[0]] if values_2d else []
        return ParseResult(
            raw_values=values_2d,
            raw_headers=raw_headers,
            detected=detected,
            needs_manual_selection=True,
        )

    # Step 2: check for multiline header
    headers, consumed_next = _check_multiline_header(values_2d, header_idx)
    data_start = header_idx + (2 if consumed_next else 1)
    detected.header_row = header_idx

    # Step 3: fuzzy-match columns
    name_idx, name_score = _fuzzy_find_col(headers, _NAME_SYNONYMS, threshold=70)
    qty_idx, qty_score = _fuzzy_find_col(headers, _QTY_SYNONYMS, threshold=70)
    code_idx, code_score = _fuzzy_find_col(headers, _CODE_SYNONYMS, threshold=70)

    detected.fuzzy_scores = {"name": name_score, "qty": qty_score, "code": code_score}
    detected.method = "fuzzy"

    # Step 4: content heuristic fallbacks
    assigned = {i for i in (name_idx, qty_idx, code_idx) if i is not None}

    if qty_idx is None:
        qty_idx = _find_qty_heuristic(values_2d, data_start, num_cols, exclude_cols=assigned)
        if qty_idx is not None:
            assigned.add(qty_idx)
            detected.method = "heuristic"

    if name_idx is None:
        name_idx = _find_name_heuristic(values_2d, data_start, num_cols, exclude_cols=assigned)
        if name_idx is not None:
            assigned.add(name_idx)
            detected.method = "heuristic"

    detected.name_idx = name_idx
    detected.qty_idx = qty_idx
    detected.code_idx = code_idx

    # Step 4c: detect combined qty+uom column
    qty_uom_combined = False
    if qty_idx is not None:
        if _combined_ratio(values_2d, data_start, qty_idx) >= 0.5:
            qty_uom_combined = True
    else:
        # No qty column found — try to find a combined qty+uom column
        combined_idx = _find_qty_uom_combined_heuristic(
            values_2d, data_start, num_cols, exclude_cols=assigned
        )
        if combined_idx is not None:
            qty_idx = combined_idx
            qty_uom_combined = True
            assigned.add(qty_idx)
            detected.qty_idx = qty_idx
            detected.method = "heuristic"
    detected.qty_uom_combined = qty_uom_combined

    # Step 4b: detect extra columns (standard, strength, note, uom)
    assigned = {i for i in (name_idx, qty_idx, code_idx) if i is not None}

    standard_idx, _ = _fuzzy_find_col(headers, _STANDARD_SYNONYMS, threshold=70)
    if standard_idx in assigned:
        standard_idx = None
    if standard_idx is not None:
        assigned.add(standard_idx)

    strength_col_idx, _ = _fuzzy_find_col(headers, _STRENGTH_COL_SYNONYMS, threshold=70)
    if strength_col_idx in assigned or strength_col_idx == standard_idx:
        strength_col_idx = None
    if strength_col_idx is not None:
        assigned.add(strength_col_idx)

    note_idx, _ = _fuzzy_find_col(headers, _NOTE_SYNONYMS, threshold=70)
    if note_idx in assigned or note_idx in (standard_idx, strength_col_idx):
        note_idx = None
    if note_idx is not None:
        assigned.add(note_idx)

    uom_idx, _ = _fuzzy_find_col(headers, _UOM_SYNONYMS, threshold=70)
    if uom_idx in assigned:
        uom_idx = None

    detected.standard_idx = standard_idx
    detected.strength_col_idx = strength_col_idx
    detected.note_idx = note_idx
    detected.uom_idx = uom_idx
    detected.low_confidence = score < 3  # score 2 = minimum passing threshold

    # Step 5: decide if we have enough
    if name_idx is None:
        from app.column_scorer import run_column_scorer  # noqa: PLC0415
        sr = run_column_scorer(values_2d, data_start=data_start)
        detected.name_idx = sr.name_idx
        detected.low_confidence = True
        detected.reasons = sr.reasons
        raw_headers = [str(v) if v is not None else "" for v in values_2d[header_idx]]
        return ParseResult(
            raw_values=values_2d,
            raw_headers=raw_headers,
            detected=detected,
            needs_manual_selection=True,
        )

    # NAME is found — build DataFrame (QTY may be None, that's ok)
    effective_header = header_idx if not consumed_next else header_idx + 1
    try:
        df = build_dataframe_from_columns(
            values_2d, effective_header, name_idx, qty_idx, code_idx,
            standard_idx=standard_idx, strength_col_idx=strength_col_idx, note_idx=note_idx,
            qty_is_combined=qty_uom_combined, uom_idx=uom_idx,
        )
    except ParseError:
        raw_headers = [str(v) if v is not None else "" for v in values_2d[header_idx]]
        return ParseResult(
            raw_values=values_2d,
            raw_headers=raw_headers,
            detected=detected,
            needs_manual_selection=True,
        )

    return ParseResult(
        df=df,
        detected=detected,
        needs_manual_selection=False,
    )


def load_excel(file_path: str | Path) -> pd.DataFrame:
    """Backward-compatible wrapper. Raises ParseError on failure."""
    result = parse_excel(file_path)
    if result.needs_manual_selection or result.df is None:
        raise ParseError(_PARSE_ERR)
    return result.df


# ── Display helpers (unchanged) ───────────────────────────────


def dataframe_preview(df: pd.DataFrame, limit: int = 200) -> pd.DataFrame:
    """Return the first `limit` rows of the DataFrame."""
    return df.head(limit)


def dataframe_to_html(df: pd.DataFrame) -> str:
    """Convert DataFrame to an HTML table string with Russian column headers."""
    from app.display_labels import display_label, format_qty

    display_df = df.copy()
    if "qty" in display_df.columns:
        display_df["qty"] = display_df["qty"].apply(format_qty)

    renamed = display_df.rename(columns=display_label)
    return renamed.to_html(
        index=False,
        classes="table",
        border=0,
        na_rep="",
    )


def dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Export DataFrame to .xlsx bytes with bold header and Russian column headers."""
    from app.display_labels import display_label

    export_df = df.rename(columns=display_label)

    # Write whole-number qty values as int (avoids "64.0" in Excel cells)
    qty_label = display_label("qty")
    if qty_label in export_df.columns:
        export_df = export_df.copy()
        export_df[qty_label] = export_df[qty_label].apply(
            lambda v: int(v) if pd.notna(v) and isinstance(v, float) and v == int(v) else v
        )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Результат")
        ws = writer.sheets["Результат"]

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        for col_idx, col_name in enumerate(export_df.columns, start=1):
            width = min(max(len(str(col_name)) + 2, 12), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    return buf.getvalue()
